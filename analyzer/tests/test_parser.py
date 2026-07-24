from io import BytesIO

from django.test import SimpleTestCase
from openpyxl import Workbook

from analyzer.services.excel_parser import parse_lotofacil_excel


def _build_excel_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    headers = ['Concurso', 'Data Sorteio'] + [f'Bola{i}' for i in range(1, 16)]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    output.name = 'lotofacil.xlsx'
    return output


class ExcelParserTests(SimpleTestCase):
    def test_parse_valid_rows(self):
        rows = [
            [1, '01/01/2024', *range(1, 16)],
            [2, '03/01/2024', *range(2, 17)],
        ]
        excel_file = _build_excel_bytes(rows)

        result = parse_lotofacil_excel(excel_file)

        self.assertEqual(result.valid_rows, 2)
        self.assertEqual(result.invalid_rows, 0)
        self.assertEqual(result.draws[0].concurso, 1)
        self.assertEqual(result.draws[1].dezenas[0], 2)

    def test_ignore_invalid_row(self):
        valid_row = [10, '01/01/2024', *range(1, 16)]
        invalid_row = [11, '02/01/2024', 1, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
        excel_file = _build_excel_bytes([valid_row, invalid_row])

        result = parse_lotofacil_excel(excel_file)

        self.assertEqual(result.valid_rows, 1)
        self.assertEqual(result.invalid_rows, 1)
        self.assertTrue(result.messages)

    def test_accepts_ball_headers_with_space(self):
        workbook = Workbook()
        sheet = workbook.active
        headers = ['Concurso', 'Data Sorteio'] + [f'Bola {i}' for i in range(1, 16)]
        sheet.append(headers)
        sheet.append([1, '01/01/2024', *range(1, 16)])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        output.name = 'lotofacil.xlsx'

        result = parse_lotofacil_excel(output)

        self.assertEqual(result.valid_rows, 1)
        self.assertEqual(result.draws[0].dezenas[0], 1)

    def test_uses_best_sheet_when_active_sheet_has_no_valid_rows(self):
        workbook = Workbook()

        sheet1 = workbook.active
        sheet1.title = 'Resumo'
        sheet1.append(['Texto'])
        sheet1.append(['Sem dados válidos'])

        sheet2 = workbook.create_sheet('Resultados')
        headers = ['Concurso', 'Data Sorteio'] + [f'Bola{i}' for i in range(1, 16)]
        sheet2.append(headers)
        sheet2.append([100, '01/01/2024', *range(1, 16)])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        output.name = 'lotofacil.xlsx'

        result = parse_lotofacil_excel(output)

        self.assertEqual(result.valid_rows, 1)
        self.assertEqual(result.draws[0].concurso, 100)
