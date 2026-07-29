import datetime as dt
import re
import unicodedata
from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass
class DrawRecord:
    concurso: int | None
    data_sorteio: dt.date | None
    dezenas: list[int]


@dataclass
class ParseResult:
    draws: list[DrawRecord]
    total_rows: int
    valid_rows: int
    invalid_rows: int
    messages: list[str]


def _normalize_header(value: object) -> str:
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r'\s+', ' ', text)
    return text


def _extract_ball_number(header_name: str) -> int | None:
    match = re.match(r'^bola\s*(\d+)$', header_name)
    if not match:
        return None

    number = int(match.group(1))
    if 1 <= number <= 15:
        return number
    return None


def _to_int(value: object) -> int | None:
    if value is None or value == '':
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None

    text = str(value).strip().replace(',', '.')
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None

    if not number.is_integer():
        return None
    return int(number)


def _to_date(value: object) -> dt.date | None:
    if value is None or value == '':
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for date_format in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return dt.datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def _find_header_row(rows: list[tuple[object, ...]], max_lines: int = 80) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows[:max_lines]):
        normalized_map: dict[str, int] = {}
        ball_columns: dict[int, int] = {}
        for col_idx, value in enumerate(row):
            header_name = _normalize_header(value)
            if header_name:
                normalized_map[header_name] = col_idx

                ball_number = _extract_ball_number(header_name)
                if ball_number is not None:
                    ball_columns[ball_number] = col_idx

        has_concurso = 'concurso' in normalized_map
        has_all_balls = set(ball_columns.keys()) == set(range(1, 16))
        if has_concurso and has_all_balls:
            for number, col_idx in ball_columns.items():
                normalized_map[f'bola{number}'] = col_idx
            return idx, normalized_map

    raise ValueError('Nao foi possivel identificar o cabecalho esperado (Concurso, Bola1..Bola15).')


def parse_lotofacil_excel(uploaded_file) -> ParseResult:
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    required_number_columns = [f'bola{i}' for i in range(1, 16)]
    sheet_attempts: list[dict] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            # Some spreadsheets are saved with incorrect dimension metadata
            # (e.g., A1:AG1) even when they contain data rows.
            worksheet.reset_dimensions()
            rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            sheet_attempts.append({'sheet': worksheet.title, 'reason': 'aba vazia'})
            continue

        try:
            header_row_index, header_map = _find_header_row(rows)
        except ValueError:
            sheet_attempts.append({'sheet': worksheet.title, 'reason': 'cabecalho nao encontrado'})
            continue

        missing = [name for name in required_number_columns if name not in header_map]
        if missing:
            sheet_attempts.append({'sheet': worksheet.title, 'reason': f'colunas ausentes: {", ".join(missing)}'})
            continue

        draws: list[DrawRecord] = []
        messages: list[str] = []
        invalid_rows = 0
        total_rows = 0

        concurso_index = header_map.get('concurso')
        data_index = header_map.get('data sorteio')

        for excel_row_number, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
            dezenas: list[int] = []
            for column_name in required_number_columns:
                number_idx = header_map[column_name]
                cell_value = row[number_idx] if number_idx < len(row) else None
                parsed_number = _to_int(cell_value)
                if parsed_number is None:
                    dezenas = []
                    break
                dezenas.append(parsed_number)

            if not dezenas:
                continue

            total_rows += 1

            dezenas_set = set(dezenas)
            if len(dezenas) != 15 or len(dezenas_set) != 15:
                invalid_rows += 1
                messages.append(f'Aba {worksheet.title}, linha {excel_row_number}: dezenas duplicadas ou incompletas.')
                continue

            if any(number < 1 or number > 25 for number in dezenas):
                invalid_rows += 1
                messages.append(f'Aba {worksheet.title}, linha {excel_row_number}: dezenas fora do intervalo 1-25.')
                continue

            concurso = None
            if concurso_index is not None and concurso_index < len(row):
                concurso = _to_int(row[concurso_index])

            data_sorteio = None
            if data_index is not None and data_index < len(row):
                data_sorteio = _to_date(row[data_index])

            draws.append(
                DrawRecord(
                    concurso=concurso,
                    data_sorteio=data_sorteio,
                    dezenas=sorted(dezenas),
                )
            )

        sheet_attempts.append(
            {
                'sheet': worksheet.title,
                'draws': draws,
                'total_rows': total_rows,
                'invalid_rows': invalid_rows,
                'messages': messages,
                'reason': (
                    'nenhuma linha com 15 dezenas encontrada apos o cabecalho'
                    if total_rows == 0
                    else 'linhas encontradas, mas todas invalidas'
                ),
            }
        )

    candidates = [item for item in sheet_attempts if item.get('draws')]
    if not candidates:
        details = ', '.join(
            f"{item['sheet']}: {item.get('reason', 'sem concursos válidos')}"
            for item in sheet_attempts
        )
        raise ValueError(f'Nenhum concurso válido encontrado na planilha. Detalhes: {details}')

    best = max(candidates, key=lambda item: (len(item['draws']), item['total_rows'], -item['invalid_rows']))

    draws = _sort_draws(best['draws'])
    valid_rows = len(draws)
    invalid_rows = best['invalid_rows']
    messages = best['messages']

    if invalid_rows > 0:
        messages.insert(0, f'{invalid_rows} linha(s) inválidas foram ignoradas.')

    if len(candidates) > 1:
        messages.insert(0, f"Aba selecionada para análise: {best['sheet']}.")

    return ParseResult(
        draws=draws,
        total_rows=best['total_rows'],
        valid_rows=valid_rows,
        invalid_rows=invalid_rows,
        messages=messages[:20],
    )


def _sort_draws(draws: list[DrawRecord]) -> list[DrawRecord]:
    has_concurso = any(item.concurso is not None for item in draws)
    if has_concurso:
        return sorted(draws, key=lambda item: (item.concurso is None, item.concurso or 0))

    has_data = any(item.data_sorteio is not None for item in draws)
    if has_data:
        return sorted(draws, key=lambda item: (item.data_sorteio is None, item.data_sorteio or dt.date.min))

    return draws
